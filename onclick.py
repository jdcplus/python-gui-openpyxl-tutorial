from tkinter import *
import openpyxl as px

class GUI(Frame):

    def __init__(self,master=None):
        Frame.__init__(self, master)
        self.grid()


        # start
        global wb 
        wb = px.load_workbook("test.xlsx")
        ws = wb.active

        grid_idx_x = 0
        grid_idx_y = 0
        self.fnameLabel = Label(master, fg = "Black",text="Printing table from Excel.")
        self.fnameLabel.grid()

        mx_rows = ws.max_row
        mx_column = ws.max_column
        INF = "9999-12-31 00:00:00"
        row_red = [False] * (mx_rows+1)
        row_green = [False] * (mx_rows+1)

        for i in range(1,mx_rows+1):
            for j in range(1,mx_column):
                if(j is 2):
                    cur = ws.cell(row=i, column=j).value
                    nxt = ws.cell(row=i, column=j+1).value
                    str_v = str(nxt)
                    #print("test.:",str_v[0:4])
                    if(cur == nxt):
                        row_red[i] = True
                    elif(str_v[0:4] == "9999"):
                        row_green[i] = True

        for i in range(1,mx_rows+1):

            grid_idx_y += 1
            grid_idx_x = 0

            for j in range(1,mx_column+1):
                val = ws.cell(row=i, column=j).value
                if(row_red[i]): # print as red
                    self.fnameLabel = Label(master, fg = "red", text='%s'%(val),borderwidth=1, width=25).grid(row=grid_idx_y,column=grid_idx_x)
                elif(row_green[i]): # print as green
                    self.fnameLabel = Label(master, fg = "green", text='%s'%(val),borderwidth=1, width=25).grid(row=grid_idx_y,column=grid_idx_x)
                else:
                    self.fnameLabel = Label(master, fg = "black", text='%s'%(val),borderwidth=1, width=25).grid(row=grid_idx_y,column=grid_idx_x)
                grid_idx_x += 1

        grid_idx_y += 1
        grid_idx_x = 1
        self.submitButton = Button(master, text="Export to Excel", command= self.excel_export, width=25).grid(row=grid_idx_y,column=grid_idx_x)
        
        #end
        '''
        self.fnameLabel = Label(master, text="First Name")
        self.fnameLabel.grid()

        self.fnameEntry = Entry(master)
        self.fnameEntry.grid()

        self.lnameLabel = Label(master, text="Last Name")
        self.lnameLabel.grid()

        self.lnameEntry = Entry(master)
        self.lnameEntry.grid()

        self.submitButton = Button(master, command=self.buttonClick, text="Submit")
        self.submitButton.grid()
        '''

    def excel_export(self):
        ws = wb.active
        wb.save('final.xlsx')
        print("file exported successfully")

if __name__ == "__main__":
    guiFrame = GUI()
    guiFrame.mainloop()
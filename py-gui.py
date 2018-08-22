from tkinter import *
import openpyxl as px

def excel_export(master):
    global export_flag
    export_flag = True
    
master = Tk()
export_flag = False
wb = px.load_workbook("test.xlsx")
ws = wb.active

grid_idx_x = 0
grid_idx_y = 0
Label(master, fg = "Black",text="Printing table from Excel.").grid(row=grid_idx_y,column=grid_idx_x)
mx_rows = ws.max_row
mx_column = ws.max_column
INF = "9999-12-31 00:00:00"
row_red = [False] * (mx_rows+1)
row_green = [False] * (mx_rows+1)

for i in range(1,mx_rows+1):
    for j in range(1,mx_column):
        if(j is 2):
            cur = ws.cell(row=i, column=j).value
            nxt = ws.cell(row=i, column=j+1).value
            str_v = str(nxt)
            #print("test.:",str_v[0:4])
            if(cur == nxt):
                row_red[i] = True
            elif(str_v[0:4] == "9999"):
                row_green[i] = True

for i in range(1,mx_rows+1):

    grid_idx_y += 1
    grid_idx_x = 0

    for j in range(1,mx_column+1):
        val = ws.cell(row=i, column=j).value
        if(row_red[i]): # print as red
            Label(master, fg = "red", text='%s'%(val),borderwidth=1, width=25).grid(row=grid_idx_y,column=grid_idx_x)
        elif(row_green[i]): # print as green
            Label(master, fg = "green", text='%s'%(val),borderwidth=1, width=25).grid(row=grid_idx_y,column=grid_idx_x)
        else:
            Label(master, fg = "black", text='%s'%(val),borderwidth=1, width=25).grid(row=grid_idx_y,column=grid_idx_x)
        grid_idx_x += 1

grid_idx_y += 1
grid_idx_x = 1
Button(master, text="Export to Excel", command= lambda: excel_export(master), width=25).grid(row=grid_idx_y,column=grid_idx_x)

if(export_flag is True):
    wb.save('final.xlsx')
    grid_idx_x = 1
    print("file exported successfully")
    Lebel(master, text="Export to Excel Successfully").grid(row=grid_idx_y,column=grid_idx_x)
else:   
    print("file discarded successfully")


wb.close()
mainloop()
